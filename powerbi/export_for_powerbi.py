"""
export_for_powerbi.py

Builds the SQLite database (if needed) then exports five clean, pre-aggregated
tables into a single Excel workbook: powerbi/RecruitmentFunnel_PowerBI.xlsx

Each sheet becomes one table in the Power BI data model. Relationships are
defined in the companion build guide.

Run from the repo root:
    python powerbi/export_for_powerbi.py
"""

import os
import sys
import sqlite3
import subprocess
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(ROOT, "database", "recruitment.db")
OUT_PATH = os.path.join(ROOT, "powerbi", "RecruitmentFunnel_PowerBI.xlsx")

CHANNEL_COST = {
    "LinkedIn": 45,
    "Referral": 10,
    "Job Board": 60,
    "Agency Database": 15,
    "Cold Outreach": 25,
}


def ensure_database():
    if not os.path.exists(DB_PATH):
        print("Database not found — generating now...")
        gen = os.path.join(ROOT, "data", "generate_data.py")
        subprocess.run([sys.executable, gen], check=True, cwd=ROOT)
    print(f"Using database: {DB_PATH}")


def get_conn():
    return sqlite3.connect(DB_PATH)


# ---------------------------------------------------------------------------
# Table 1: Candidates  (raw dimension — cleaned source_channel)
# ---------------------------------------------------------------------------

def build_candidates(conn):
    df = pd.read_sql("""
        SELECT
            candidate_id,
            TRIM(UPPER(SUBSTR(TRIM(source_channel), 1, 1)) ||
                 LOWER(SUBSTR(TRIM(source_channel), 2)))  AS source_channel,
            role_applied,
            years_experience,
            expected_salary,
            date_sourced
        FROM candidates
    """, conn)
    df["date_sourced"] = pd.to_datetime(df["date_sourced"])
    return df


# ---------------------------------------------------------------------------
# Table 2: Pipeline_Stages  (fact table)
# ---------------------------------------------------------------------------

def build_pipeline(conn):
    df = pd.read_sql("""
        SELECT
            ps.stage_id,
            ps.candidate_id,
            ps.role_id,
            ps.recruiter,
            ps.stage,
            ps.stage_date,
            TRIM(UPPER(SUBSTR(TRIM(c.source_channel), 1, 1)) ||
                 LOWER(SUBSTR(TRIM(c.source_channel), 2))) AS source_channel
        FROM pipeline_stages ps
        JOIN candidates c ON c.candidate_id = ps.candidate_id
    """, conn)
    df["stage_date"] = pd.to_datetime(df["stage_date"])
    # Numeric stage order for funnel sorting in Power BI
    order = {"Sourced": 1, "Screened": 2, "Interviewed": 3,
             "Offered": 4, "Placed": 5, "Rejected": 6}
    df["stage_order"] = df["stage"].map(order)
    return df


# ---------------------------------------------------------------------------
# Table 3: Channel_ROI  (pre-aggregated — drives the ROI visuals)
# ---------------------------------------------------------------------------

def build_channel_roi(conn):
    df = pd.read_sql("""
        SELECT
            TRIM(UPPER(SUBSTR(TRIM(c.source_channel), 1, 1)) ||
                 LOWER(SUBSTR(TRIM(c.source_channel), 2))) AS source_channel,
            COUNT(DISTINCT c.candidate_id)               AS total_sourced,
            SUM(CASE WHEN ps.stage = 'Screened'    THEN 1 ELSE 0 END) AS screened,
            SUM(CASE WHEN ps.stage = 'Interviewed' THEN 1 ELSE 0 END) AS interviewed,
            SUM(CASE WHEN ps.stage = 'Offered'     THEN 1 ELSE 0 END) AS offered,
            SUM(CASE WHEN ps.stage = 'Placed'      THEN 1 ELSE 0 END) AS placed
        FROM candidates c
        JOIN pipeline_stages ps ON ps.candidate_id = c.candidate_id
        GROUP BY 1
    """, conn)

    df["cost_per_sourced"] = df["source_channel"].map(CHANNEL_COST)
    df["total_sourcing_cost"] = df["total_sourced"] * df["cost_per_sourced"]
    df["cost_per_hire"] = (df["total_sourcing_cost"] / df["placed"]).round(2)
    df["screened_rate"]    = (df["screened"]    / df["total_sourced"] * 100).round(1)
    df["interviewed_rate"] = (df["interviewed"] / df["total_sourced"] * 100).round(1)
    df["offered_rate"]     = (df["offered"]     / df["total_sourced"] * 100).round(1)
    df["placement_rate"]   = (df["placed"]      / df["total_sourced"] * 100).round(1)
    return df


# ---------------------------------------------------------------------------
# Table 4: Funnel_Summary  (stage-level aggregation for funnel chart)
# ---------------------------------------------------------------------------

def build_funnel_summary(conn):
    df = pd.read_sql("""
        SELECT
            stage,
            COUNT(*) AS candidate_count
        FROM pipeline_stages
        WHERE stage != 'Rejected'
        GROUP BY stage
    """, conn)
    order = {"Sourced": 1, "Screened": 2, "Interviewed": 3, "Offered": 4, "Placed": 5}
    df["stage_order"] = df["stage"].map(order)
    df = df.sort_values("stage_order")
    df["drop_off_pct"] = (
        (1 - df["candidate_count"] / df["candidate_count"].shift(1)) * 100
    ).round(1)
    return df


# ---------------------------------------------------------------------------
# Table 5: Recruiter_Performance  (recruiter leaderboard)
# ---------------------------------------------------------------------------

def build_recruiter_perf(conn):
    df = pd.read_sql("""
        SELECT
            ps.recruiter,
            COUNT(DISTINCT ps.candidate_id)                          AS total_candidates,
            SUM(CASE WHEN ps.stage = 'Placed' THEN 1 ELSE 0 END)    AS placements,
            SUM(CASE WHEN ps.stage = 'Screened' THEN 1 ELSE 0 END)  AS screened,
            SUM(CASE WHEN ps.stage = 'Interviewed' THEN 1 ELSE 0 END) AS interviewed
        FROM pipeline_stages ps
        GROUP BY ps.recruiter
    """, conn)
    df["placement_rate_pct"] = (df["placements"] / df["total_candidates"] * 100).round(1)

    # Avg time-to-fill: days from Sourced to Placed per role
    ttf = pd.read_sql("""
        SELECT
            ps.recruiter,
            ps.role_id,
            MIN(CASE WHEN ps.stage = 'Sourced' THEN ps.stage_date END) AS sourced_date,
            MAX(CASE WHEN ps.stage = 'Placed'  THEN ps.stage_date END) AS placed_date
        FROM pipeline_stages ps
        GROUP BY ps.recruiter, ps.role_id
        HAVING placed_date IS NOT NULL
    """, conn)
    ttf["sourced_date"] = pd.to_datetime(ttf["sourced_date"])
    ttf["placed_date"]  = pd.to_datetime(ttf["placed_date"])
    ttf["days_to_fill"] = (ttf["placed_date"] - ttf["sourced_date"]).dt.days
    avg_ttf = ttf.groupby("recruiter")["days_to_fill"].mean().round(1).reset_index()
    avg_ttf.columns = ["recruiter", "avg_days_to_fill"]

    df = df.merge(avg_ttf, on="recruiter", how="left")
    return df


# ---------------------------------------------------------------------------
# Table 6: Monthly_Trends  (time series for line charts)
# ---------------------------------------------------------------------------

def build_monthly_trends(conn):
    df = pd.read_sql("""
        SELECT
            strftime('%Y-%m', stage_date) AS year_month,
            stage,
            COUNT(*) AS count
        FROM pipeline_stages
        WHERE stage IN ('Sourced', 'Placed')
        GROUP BY 1, 2
        ORDER BY 1, 2
    """, conn)
    df["year_month"] = pd.to_datetime(df["year_month"] + "-01")
    return df.pivot(index="year_month", columns="stage", values="count").reset_index().fillna(0)


# ---------------------------------------------------------------------------
# Excel writer with styled headers
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def style_sheet(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)


def write_excel(sheets: dict):
    with pd.ExcelWriter(OUT_PATH, engine="openpyxl", datetime_format="YYYY-MM-DD") as w:
        for name, df in sheets.items():
            df.to_excel(w, sheet_name=name, index=False)
    wb = load_workbook(OUT_PATH)
    for ws in wb.worksheets:
        style_sheet(ws)
    wb.save(OUT_PATH)
    print(f"Saved: {OUT_PATH}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    ensure_database()
    conn = get_conn()

    sheets = {
        "Candidates":           build_candidates(conn),
        "Pipeline_Stages":      build_pipeline(conn),
        "Channel_ROI":          build_channel_roi(conn),
        "Funnel_Summary":       build_funnel_summary(conn),
        "Recruiter_Performance": build_recruiter_perf(conn),
        "Monthly_Trends":       build_monthly_trends(conn),
    }

    conn.close()
    write_excel(sheets)

    print("\nSheets written:")
    for name, df in sheets.items():
        print(f"  {name}: {len(df)} rows x {len(df.columns)} cols")


if __name__ == "__main__":
    main()
